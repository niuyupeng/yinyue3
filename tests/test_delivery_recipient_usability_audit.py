from __future__ import annotations

import csv
import zipfile
from pathlib import Path

from openpyxl import Workbook

from chorale.delivery_recipient_usability_audit import (
    EXPECTED_WORKBOOK_HEADERS,
    ISSUE_TEMPLATE_HEADERS,
    REQUIRED_FORMS,
    REQUIRED_PLAYBACK_VARIANTS,
    REQUIRED_TOP_LEVEL,
    audit_recipient_usability,
)


def _write_package(root: Path, *, missing_mp3: bool = False, bad_workbook: bool = False) -> None:
    for rel in REQUIRED_TOP_LEVEL:
        path = root / rel
        path.parent.mkdir(parents=True, exist_ok=True)
        text = "Project1 recipient file"
        if rel == "START_HERE_CN.html":
            text = "交付包入口 score_audio_player.html forms/project1_expert_rating_forms_CN.xlsx MusicXML 不是神经音频生成"
        elif rel == "RETURN_FILES_CHECKLIST.md":
            text = (
                "project1_expert_rating_forms_CN.xlsx "
                "rater_background_form_project1_CN.csv "
                "absolute_rating_form_project1_CN.csv "
                "paired_comparison_form_project1_CN.csv"
            )
        elif rel == "OPEN_PROJECT1_REVIEW_PACKAGE.ps1":
            text = "PROJECT1_PACKAGE_SELF_TEST.ps1 START_HERE_CN.html score_audio_player.html"
        elif rel == "PROJECT1_PACKAGE_SELF_TEST.ps1":
            text = "project1_recipient_package_self_test_v1 Project1 package self-test"
        elif rel == "VERIFY_DELIVERY_INTEGRITY.ps1":
            text = "DELIVERY_FILE_MANIFEST SHA256"
        path.write_text(text, encoding="utf-8")

    for rel in REQUIRED_FORMS:
        if rel.endswith(".xlsx"):
            _write_workbook(root / rel, bad=bad_workbook)
        else:
            path = root / rel
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text("中文表格说明", encoding="utf-8")

    _write_issue_template(root / "REVIEW_ISSUE_REPORT_TEMPLATE.csv")
    _write_correspondence(root, missing_mp3=missing_mp3)


def _write_workbook(path: Path, *, bad: bool) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers in EXPECTED_WORKBOOK_HEADERS.items():
        sheet = workbook.create_sheet(sheet_name)
        values = list(headers)
        if bad and sheet_name == "逐首评分":
            values[1] = "错误列名"
        for column, value in enumerate(values, start=1):
            sheet.cell(row=1, column=column, value=value)
    workbook.save(path)


def _write_issue_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(ISSUE_TEMPLATE_HEADERS)
        writer.writerow(["ISSUE001", "P1S01", "absolute", "full_choir", "12.5", "score_audio", "3", "", "yes", "", ""])


def _write_correspondence(root: Path, *, missing_mp3: bool) -> None:
    fieldnames = [
        "group",
        "score_id",
        "variant",
        "source_musicxml",
        "render_musicxml",
        "midi",
        "wav",
        "mp3",
        "status",
    ]
    rows = []
    variants = sorted(REQUIRED_PLAYBACK_VARIANTS)
    for idx in range(40):
        group = "absolute" if idx < 20 else "paired"
        score_id = f"P1S{idx + 1:02d}"
        for variant in variants:
            source = f"{group}_score_musicxml/{score_id}.musicxml"
            render = f"render_xml/{group}/{score_id}/{score_id}_{variant}.musicxml"
            midi = f"midi_pro/{group}/{score_id}/{score_id}_{variant}.mid"
            mp3 = f"audio_pro/{group}/{score_id}/{score_id}_{variant}.mp3"
            for rel in [source, render, midi]:
                (root / rel).parent.mkdir(parents=True, exist_ok=True)
                (root / rel).write_bytes(b"x")
            if not (missing_mp3 and score_id == "P1S01" and variant == "stem_alto"):
                (root / mp3).parent.mkdir(parents=True, exist_ok=True)
                (root / mp3).write_bytes(b"ID3")
            rows.append(
                {
                    "group": group,
                    "score_id": score_id,
                    "variant": variant,
                    "source_musicxml": source,
                    "render_musicxml": render,
                    "midi": midi,
                    "wav": "",
                    "mp3": mp3,
                    "status": "ok",
                }
            )
    with (root / "SCORE_AUDIO_CORRESPONDENCE.csv").open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def test_recipient_usability_audit_passes_complete_package(tmp_path: Path) -> None:
    _write_package(tmp_path)

    summary = audit_recipient_usability(package_dir=tmp_path)

    assert summary["all_pass"] is True
    assert summary["recipient_usability_score"] == 100
    assert summary["score_audio_correspondence"]["rows"] == 240


def test_recipient_usability_audit_detects_bad_workbook_header(tmp_path: Path) -> None:
    _write_package(tmp_path, bad_workbook=True)

    summary = audit_recipient_usability(package_dir=tmp_path)

    assert summary["all_pass"] is False
    assert any("workbook sheet" in issue for issue in summary["issues"])


def test_recipient_usability_audit_detects_missing_score_audio_reference(tmp_path: Path) -> None:
    _write_package(tmp_path, missing_mp3=True)

    summary = audit_recipient_usability(package_dir=tmp_path)

    assert summary["all_pass"] is False
    assert any("missing references" in issue for issue in summary["issues"])


def test_recipient_usability_audit_reads_zip(tmp_path: Path) -> None:
    package = tmp_path / "package"
    _write_package(package)
    zip_path = tmp_path / "package.zip"
    with zipfile.ZipFile(zip_path, "w") as archive:
        for item in package.rglob("*"):
            if item.is_file():
                archive.write(item, item.relative_to(package).as_posix())

    summary = audit_recipient_usability(zip_file=zip_path)

    assert summary["all_pass"] is True
    assert summary["recipient_usability_score"] == 100
