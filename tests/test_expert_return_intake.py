from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from chorale.expert_eval_tools import write_expert_xlsx_forms
from chorale.expert_return_intake import validate_returned_ratings


def test_expert_return_intake_pending_when_no_files(tmp_path: Path) -> None:
    report = validate_returned_ratings(tmp_path / "missing")

    assert report["status"] == "expert evaluation pending"
    assert report["rating_file_count"] == 0


def test_expert_return_intake_validates_single_workbook_but_waits_for_three_raters(tmp_path: Path) -> None:
    package = make_package(tmp_path)
    xlsx = write_expert_xlsx_forms(package)
    wb = load_workbook(xlsx)
    fill_complete_workbook(wb, "R01")
    returned = tmp_path / "returned"
    returned.mkdir()
    out = returned / "rater01.xlsx"
    wb.save(out)

    report = validate_returned_ratings(returned)

    assert report["status"] == "expert evaluation pending"
    assert report["valid_rating_file_count"] == 1
    assert report["absolute_completed_rows"] == 1
    assert report["paired_completed_rows"] == 1
    assert "need at least 3 valid returned workbooks" in str(report["release_gate_issues"])


def test_expert_return_intake_ready_after_three_distinct_complete_raters(tmp_path: Path) -> None:
    package = make_package(tmp_path)
    xlsx = write_expert_xlsx_forms(package)
    returned = tmp_path / "returned"
    returned.mkdir()
    for idx in range(3):
        wb = load_workbook(xlsx)
        fill_complete_workbook(wb, f"R{idx + 1:02d}")
        wb.save(returned / f"rater{idx + 1:02d}.xlsx")

    report = validate_returned_ratings(returned)

    assert report["status"] == "ready_to_summarize"
    assert report["ready_for_commercial_summary"] is True
    assert report["valid_rating_file_count"] == 3
    assert report["absolute_completed_rows"] == 3
    assert report["paired_completed_rows"] == 3
    assert report["duplicate_rater_ids"] == []


def test_expert_return_intake_rejects_out_of_range_scores(tmp_path: Path) -> None:
    package = make_package(tmp_path)
    xlsx = write_expert_xlsx_forms(package)
    wb = load_workbook(xlsx)
    fill_complete_workbook(wb, "R01")
    wb["逐首评分"]["C2"] = 8
    returned = tmp_path / "returned"
    returned.mkdir()
    out = returned / "rater01.xlsx"
    wb.save(out)

    report = validate_returned_ratings(returned)

    assert report["status"] == "expert evaluation pending"
    assert report["valid_rating_file_count"] == 0
    assert "out of range" in str(report["files"][0]["issues"])


def test_expert_return_intake_rejects_invalid_paired_choice(tmp_path: Path) -> None:
    package = make_package(tmp_path)
    xlsx = write_expert_xlsx_forms(package)
    wb = load_workbook(xlsx)
    fill_complete_workbook(wb, "R01")
    wb["AB配对比较"]["D2"] = "A"
    wb["AB配对比较"]["L2"] = 9
    returned = tmp_path / "returned"
    returned.mkdir()
    out = returned / "rater01.xlsx"
    wb.save(out)

    report = validate_returned_ratings(returned)

    assert report["status"] == "expert evaluation pending"
    assert report["valid_rating_file_count"] == 0
    issues = str(report["files"][0]["issues"])
    assert "paired preference invalid" in issues
    assert "paired confidence out of range" in issues


def test_expert_return_intake_rejects_duplicate_rater_ids(tmp_path: Path) -> None:
    package = make_package(tmp_path)
    xlsx = write_expert_xlsx_forms(package)
    returned = tmp_path / "returned"
    returned.mkdir()
    for idx in range(3):
        wb = load_workbook(xlsx)
        fill_complete_workbook(wb, "R01")
        wb.save(returned / f"rater{idx + 1:02d}.xlsx")

    report = validate_returned_ratings(returned)

    assert report["status"] == "expert evaluation pending"
    assert report["valid_rating_file_count"] == 3
    assert report["duplicate_rater_ids"] == ["R01"]


def test_expert_return_intake_rejects_incomplete_rows(tmp_path: Path) -> None:
    package = make_package(tmp_path)
    xlsx = write_expert_xlsx_forms(package)
    wb = load_workbook(xlsx)
    wb["专家背景"]["A2"] = "R01"
    wb["逐首评分"]["C2"] = 5
    wb["AB配对比较"]["D2"] = "A_slightly"
    returned = tmp_path / "returned"
    returned.mkdir()
    out = returned / "rater01.xlsx"
    wb.save(out)

    report = validate_returned_ratings(returned)

    assert report["status"] == "expert evaluation pending"
    assert report["valid_rating_file_count"] == 0
    issues = str(report["files"][0]["issues"])
    assert "absolute rating incomplete row" in issues
    assert "paired comparison incomplete row" in issues


def make_package(tmp_path: Path) -> Path:
    package = tmp_path / "package"
    (package / "forms").mkdir(parents=True)
    (package / "absolute_score_musicxml").mkdir()
    (package / "paired_comparison_musicxml").mkdir()
    (package / "absolute_score_musicxml" / "P1S01.musicxml").write_text("<score-partwise/>", encoding="utf-8")
    (package / "paired_comparison_musicxml" / "P1P01_A.musicxml").write_text("<score-partwise/>", encoding="utf-8")
    (package / "paired_comparison_musicxml" / "P1P01_B.musicxml").write_text("<score-partwise/>", encoding="utf-8")
    return package


def fill_complete_workbook(wb, rater_id: str) -> None:
    wb["专家背景"]["A2"] = rater_id
    absolute = wb["逐首评分"]
    for col in range(3, 11):
        absolute.cell(2, col).value = 4
    paired = wb["AB配对比较"]
    for col in range(4, 12):
        paired.cell(2, col).value = "A_slightly"
    paired["L2"] = 5
