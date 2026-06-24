from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from chorale.expert_eval_tools import (
    ABSOLUTE_COLUMNS,
    PAIRED_COLUMNS,
    summarize_expert_ratings,
    write_expert_xlsx_forms,
)


def test_expert_eval_forms_are_clean_chinese(tmp_path: Path) -> None:
    package = make_package(tmp_path)

    xlsx = write_expert_xlsx_forms(package)

    wb = load_workbook(xlsx)
    assert "评分说明" in wb.sheetnames
    assert "逐首评分" in wb.sheetnames
    assert "AB配对比较" in wb.sheetnames
    assert [cell.value for cell in wb["逐首评分"][1]][: len(ABSOLUTE_COLUMNS)] == ABSOLUTE_COLUMNS
    assert [cell.value for cell in wb["AB配对比较"][1]][: len(PAIRED_COLUMNS)] == PAIRED_COLUMNS
    instructions = (package / "forms" / "中文填表说明.md").read_text(encoding="utf-8")
    assert "SATB 四部合唱乐谱评价" in instructions
    assert "鐠" not in instructions


def test_expert_eval_summary_reads_clean_workbook(tmp_path: Path) -> None:
    package = make_package(tmp_path)
    xlsx = write_expert_xlsx_forms(package)
    wb = load_workbook(xlsx)
    wb["逐首评分"]["C2"] = 4
    wb["AB配对比较"]["D2"] = "A_slightly"
    wb["AB配对比较"]["L2"] = 5
    wb.save(xlsx)

    summary = summarize_expert_ratings(xlsx, tmp_path / "results")

    assert summary["status"] == "completed"
    assert summary["absolute_completed_rows"] == 1
    assert summary["paired_completed_rows"] == 1


def make_package(tmp_path: Path) -> Path:
    package = tmp_path / "package"
    (package / "forms").mkdir(parents=True)
    (package / "absolute_score_musicxml").mkdir()
    (package / "paired_comparison_musicxml").mkdir()
    (package / "absolute_score_musicxml" / "P1S01.musicxml").write_text("<score-partwise/>", encoding="utf-8")
    (package / "paired_comparison_musicxml" / "P1P01_A.musicxml").write_text("<score-partwise/>", encoding="utf-8")
    (package / "paired_comparison_musicxml" / "P1P01_B.musicxml").write_text("<score-partwise/>", encoding="utf-8")
    return package
