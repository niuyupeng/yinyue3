from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ABSOLUTE_COLUMNS = [
    "score_id",
    "乐谱文件名",
    "和声正确性_1到5",
    "声部进行与对位正确性_1到5",
    "七和弦解决正确性_1到5",
    "终止式质量_1到5",
    "SATB可唱性_1到5",
    "传统众赞歌风格一致性_1到5",
    "作曲和声教学用途_1到5",
    "整体质量_1到5",
    "发现的主要问题",
    "备注",
]
PAIRED_COLUMNS = [
    "pair_id",
    "版本A文件",
    "版本B文件",
    "和声方面更好者",
    "声部进行方面更好者",
    "七和弦解决方面更好者",
    "终止式方面更好者",
    "SATB可唱性更好者",
    "风格一致性更好者",
    "教学用途更好者",
    "总体更好者",
    "判断信心_1到5",
    "备注",
]
BACKGROUND_COLUMNS = [
    "评分人编号",
    "最高音乐训练",
    "主要专长",
    "和声对位经验年数",
    "合唱或作曲教学经验年数",
    "是否经常使用传统功能和声_yes_no",
    "备注",
]
RUBRIC_ROWS = [
    ["分数", "含义"],
    [5, "优秀：几乎没有明显问题，可作为较高质量 SATB 和声化示例。"],
    [4, "良好：整体可用，只有少量局部声部进行、和声或终止式问题。"],
    [3, "中等：基本成立，但存在多处需要修改的问题。"],
    [2, "较差：有明显和声、对位或可唱性问题，需要较大修改。"],
    [1, "不可用：严重违背常规写作，难以作为教学或创作参考。"],
    ["A/B", "配对比较列填写 A_strongly、A_slightly、tie、B_slightly、B_strongly 或 uncertain；不要写模型名称。"],
]

ABSOLUTE_SHEET_NAMES = [
    "逐首评分",
    "逐首乐谱评分",
    "Absolute ratings",
    "閫愰璇勫垎",
    "閫愰涔愯氨璇勫垎",
    "闁劙顩荤拠鍕瀻",
]
PAIRED_SHEET_NAMES = [
    "AB配对比较",
    "A/B配对比较",
    "Paired comparisons",
    "AB閰嶅姣旇緝",
    "A/B閰嶅姣旇緝",
    "AB闁板秴顕В鏃囩窛",
]
PREFERENCE_VALUES = ["A_strongly", "A_slightly", "tie", "B_slightly", "B_strongly", "uncertain"]


def write_expert_xlsx_forms(package_dir: str | Path) -> Path:
    package = Path(package_dir)
    forms = package / "forms"
    forms.mkdir(parents=True, exist_ok=True)
    write_clean_csv_forms(package)
    write_clean_form_instructions(forms)

    wb = Workbook()
    default = wb.active
    default.title = "评分说明"
    write_rows(default, RUBRIC_ROWS)

    background = wb.create_sheet("专家背景")
    write_rows(background, [BACKGROUND_COLUMNS, ["R01", "", "", "", "", "", ""]])

    absolute = wb.create_sheet("逐首评分")
    write_rows(absolute, load_csv_rows(forms / "absolute_rating_form_project1_CN.csv", ABSOLUTE_COLUMNS))

    paired = wb.create_sheet("AB配对比较")
    write_rows(paired, load_csv_rows(forms / "paired_comparison_form_project1_CN.csv", PAIRED_COLUMNS))

    for sheet in wb.worksheets:
        style_sheet(sheet)
    add_absolute_validations(absolute)
    add_paired_validations(paired)
    add_background_validations(background)

    path = forms / "project1_expert_rating_forms_CN.xlsx"
    wb.save(path)
    return path


def write_clean_csv_forms(package: Path) -> None:
    forms = package / "forms"
    forms.mkdir(parents=True, exist_ok=True)
    absolute_files = sorted((package / "absolute_score_musicxml").glob("*.musicxml"))
    paired_files = sorted((package / "paired_comparison_musicxml").glob("*.musicxml"))
    pair_ids = sorted({path.stem.rsplit("_", 1)[0] for path in paired_files if "_" in path.stem})

    write_csv_rows(
        forms / "rater_background_form_project1_CN.csv",
        [BACKGROUND_COLUMNS, ["R01", "", "", "", "", "", ""]],
    )
    write_csv_rows(
        forms / "absolute_rating_form_project1_CN.csv",
        [ABSOLUTE_COLUMNS]
        + [[path.stem, path.name, "", "", "", "", "", "", "", "", "", ""] for path in absolute_files],
    )
    write_csv_rows(
        forms / "paired_comparison_form_project1_CN.csv",
        [PAIRED_COLUMNS]
        + [
            [
                pair_id,
                f"{pair_id}_A.musicxml",
                f"{pair_id}_B.musicxml",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
            for pair_id in pair_ids
        ],
    )


def write_clean_form_instructions(forms: Path) -> None:
    (forms / "中文填表说明.md").write_text(
        """# Project 1 专家评分表中文说明

本评价是 SATB 四部合唱乐谱评价，不是音频制作质量评价。MP3/MIDI 只用于辅助核对音高、节奏和声部进行；最终评分以 PDF/MusicXML 乐谱为准。

## 需要填写的三个表

1. `rater_background_form_project1_CN.csv`：每位专家只填写一次，用于记录匿名背景信息。
2. `absolute_rating_form_project1_CN.csv`：逐首乐谱评分，每首匿名乐谱在各维度给 1 到 5 分。
3. `paired_comparison_form_project1_CN.csv`：A/B 配对比较，不给 A 和 B 分别打分，而是在每个维度选择哪一份更好。

推荐直接填写 `project1_expert_rating_forms_CN.xlsx`，三个表已经合并成一个工作簿，并带有下拉选项。

## 逐首评分标准

- 5 = 优秀：几乎没有明显问题，可作为较高质量 SATB 和声化示例。
- 4 = 良好：整体可用，只有少量局部声部进行、和声或终止式问题。
- 3 = 中等：基本成立，但存在多处需要修改的问题。
- 2 = 较差：有明显和声、对位或可唱性问题，需要较大修改。
- 1 = 不可用：严重违背常规写作，难以作为教学或创作参考。

## A/B 配对比较

每一组 A/B 是两份匿名乐谱。请在每个维度选择：

- `A_strongly` = A 明显更好
- `A_slightly` = A 稍好
- `tie` = 差不多
- `B_slightly` = B 稍好
- `B_strongly` = B 明显更好
- `uncertain` = 无法判断
""",
        encoding="utf-8",
    )
    (forms / "评分参考与示例_CN.md").write_text(
        """# 评分参考与示例

评分不是寻找“听起来更顺耳”的音频，而是判断书面 SATB 和声化是否符合常见调性和声、声部进行与合唱可唱性。音频只作为核对材料，不能替代看谱判断。

## 和声正确性

- 5 分：功能进行自然，转位、和弦外音和局部进行基本合理。
- 3 分：整体能够成立，但有若干不自然的和弦或进行。
- 1 分：和声逻辑混乱，频繁出现难以解释的纵向音响。

## 声部进行与对位正确性

重点看声部交叉、声部间距、平行五度/八度、跳进后的回收、倾向音解决。

## 七和弦解决正确性

若出现明显七和弦，重点看七音是否按风格向下级进解决。没有明显七和弦时不要强行扣分。

## 终止式质量

重点看乐句末尾是否形成风格可信的半终止、正格终止或其他清晰闭合。

## SATB 可唱性

看 SATB 音域、旋律跳进、持续高音/低音、声部独立性是否适合合唱。

## 示例判断

如果某份乐谱整体和声成立，但出现一处平行八度和一处不自然终止，可以考虑在“声部进行”或“终止式”维度给 3 或 4 分，而不是所有维度都低分。
""",
        encoding="utf-8",
    )


def load_csv_rows(path: Path, fallback_header: list[str]) -> list[list[str]]:
    if not path.is_file():
        return [fallback_header]
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        rows = [row for row in csv.reader(handle)]
    return rows or [fallback_header]


def write_csv_rows(path: Path, rows: list[list[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def write_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 36)


def add_absolute_validations(sheet) -> None:
    score_validation = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    sheet.add_data_validation(score_validation)
    for col in range(3, 11):
        score_validation.add(f"{sheet.cell(1, col).column_letter}2:{sheet.cell(1, col).column_letter}200")


def add_paired_validations(sheet) -> None:
    preference_validation = DataValidation(
        type="list",
        formula1='"A_strongly,A_slightly,tie,B_slightly,B_strongly,uncertain"',
        allow_blank=True,
    )
    confidence_validation = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    sheet.add_data_validation(preference_validation)
    sheet.add_data_validation(confidence_validation)
    for col in range(4, 12):
        preference_validation.add(f"{sheet.cell(1, col).column_letter}2:{sheet.cell(1, col).column_letter}200")
    confidence_validation.add("L2:L200")


def add_background_validations(sheet) -> None:
    yes_no = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    sheet.add_data_validation(yes_no)
    yes_no.add("F2:F100")


def summarize_expert_ratings(ratings_path: str | Path, out_dir: str | Path) -> dict[str, object]:
    ratings = Path(ratings_path)
    return summarize_expert_rating_files([ratings], out_dir)


def summarize_expert_rating_files(rating_paths: list[str | Path], out_dir: str | Path) -> dict[str, object]:
    paths = [Path(path) for path in rating_paths]
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    absolute_rows: list[dict[str, object]] = []
    paired_rows: list[dict[str, object]] = []
    for path in paths:
        for row in read_absolute_sheet(path):
            row["rating_file"] = str(path)
            absolute_rows.append(row)
        for row in read_paired_sheet(path):
            row["rating_file"] = str(path)
            paired_rows.append(row)
    absolute_summary = summarize_absolute(absolute_rows)
    paired_summary = summarize_paired(paired_rows)
    summary = {
        "rating_files": [str(path) for path in paths],
        "rating_file_count": len(paths),
        "absolute_completed_rows": len(absolute_rows),
        "paired_completed_rows": len(paired_rows),
        "status": "completed" if absolute_rows or paired_rows else "expert evaluation pending",
        "absolute": absolute_summary,
        "paired": paired_summary,
    }
    (out / "project1_expert_eval_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    write_summary_csv(out / "project1_expert_eval_absolute_summary.csv", absolute_summary)
    write_summary_csv(out / "project1_expert_eval_paired_summary.csv", paired_summary)
    write_latex_tables(summary, Path("paper") / "tables")
    return summary


def summarize_expert_ratings_dir(ratings_dir: str | Path, out_dir: str | Path) -> dict[str, object]:
    directory = Path(ratings_dir)
    paths = sorted(
        path
        for path in directory.glob("*.xlsx")
        if not path.name.startswith("~$") and "template" not in path.name.lower()
    )
    if not paths:
        out = Path(out_dir)
        out.mkdir(parents=True, exist_ok=True)
        summary = {
            "rating_files": [],
            "rating_file_count": 0,
            "absolute_completed_rows": 0,
            "paired_completed_rows": 0,
            "status": "expert evaluation pending",
            "absolute": summarize_absolute([]),
            "paired": summarize_paired([]),
        }
        (out / "project1_expert_eval_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
        write_summary_csv(out / "project1_expert_eval_absolute_summary.csv", summary["absolute"])
        write_summary_csv(out / "project1_expert_eval_paired_summary.csv", summary["paired"])
        write_latex_tables(summary, Path("paper") / "tables")
        return summary
    return summarize_expert_rating_files(paths, out_dir)


def read_absolute_sheet(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    sheet = select_sheet(wb, ABSOLUTE_SHEET_NAMES)
    rows = rows_as_dicts(sheet)
    metric_cols = ABSOLUTE_COLUMNS[2:10]
    return [row for row in rows if any(row.get(col) not in (None, "") for col in metric_cols)]


def read_paired_sheet(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    sheet = select_sheet(wb, PAIRED_SHEET_NAMES)
    rows = rows_as_dicts(sheet)
    preference_cols = PAIRED_COLUMNS[3:11]
    return [row for row in rows if any(row.get(col) not in (None, "") for col in preference_cols)]


def select_sheet(wb, names: list[str]):
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def rows_as_dicts(sheet) -> list[dict[str, object]]:
    header = [cell.value for cell in sheet[1]]
    rows: list[dict[str, object]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(header, values)))
    return rows


def summarize_absolute(rows: list[dict[str, object]]) -> dict[str, dict[str, float]]:
    result: dict[str, dict[str, float]] = {}
    for metric in ABSOLUTE_COLUMNS[2:10]:
        values = [float(row[metric]) for row in rows if isinstance(row.get(metric), (int, float))]
        result[metric] = {
            "n": float(len(values)),
            "mean": round(sum(values) / len(values), 3) if values else 0.0,
        }
    return result


def summarize_paired(rows: list[dict[str, object]]) -> dict[str, dict[str, float]]:
    result: dict[str, dict[str, float]] = {}
    for metric in PAIRED_COLUMNS[3:11]:
        counts = {key: 0 for key in PREFERENCE_VALUES}
        for row in rows:
            value = str(row.get(metric) or "").strip()
            if value in counts:
                counts[value] += 1
        total = sum(counts.values())
        result[metric] = {key: float(value) for key, value in counts.items()}
        result[metric]["n"] = float(total)
    return result


def write_summary_csv(path: Path, summary: dict[str, dict[str, float]]) -> None:
    keys = sorted({key for values in summary.values() for key in values})
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=["metric", *keys])
        writer.writeheader()
        for metric, values in summary.items():
            writer.writerow({"metric": metric, **values})


def write_latex_tables(summary: dict[str, object], tables_dir: Path) -> None:
    tables_dir.mkdir(parents=True, exist_ok=True)
    path = tables_dir / "project1_expert_eval_results.tex"
    if summary.get("status") != "completed":
        path.write_text(make_pending_latex_table(), encoding="utf-8")
        return
    absolute = summary.get("absolute", {})
    paired = summary.get("paired", {})
    lines = [
        r"\begin{table}[t]",
        r"\centering",
        r"\caption{Expert evaluation summary. Values are computed only from completed returned rating forms.}",
        r"\begin{tabular}{lcc}",
        r"\toprule",
        r"Absolute rating dimension & Mean & Completed ratings \\",
        r"\midrule",
    ]
    if isinstance(absolute, dict):
        for metric, values in absolute.items():
            if isinstance(values, dict):
                lines.append(
                    f"{latex_escape(metric)} & {float(values.get('mean', 0.0)):.3f} & {int(float(values.get('n', 0.0)))} \\\\"
                )
    lines.extend(
        [
            r"\midrule",
            r"Paired comparison dimension & A/B/Tie/Uncertain counts & Completed comparisons \\",
        ]
    )
    if isinstance(paired, dict):
        for metric, values in paired.items():
            if isinstance(values, dict):
                counts = "/".join(str(int(float(values.get(key, 0.0)))) for key in PREFERENCE_VALUES)
                lines.append(f"{latex_escape(metric)} & {counts} & {int(float(values.get('n', 0.0)))} \\\\")
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{table}", ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def make_pending_latex_table() -> str:
    return "\n".join(
        [
            r"\begin{table}[t]",
            r"\centering",
            r"\caption{Expert evaluation status. No completed expert rating forms have been ingested yet.}",
            r"\begin{tabular}{ll}",
            r"\toprule",
            r"Item & Status \\",
            r"\midrule",
            r"Absolute score ratings & expert evaluation pending \\",
            r"Paired A/B comparisons & expert evaluation pending \\",
            r"Expert-rating summary statistics & not available until returned forms are ingested \\",
            r"\bottomrule",
            r"\end{tabular}",
            r"\end{table}",
            "",
        ]
    )


def latex_escape(value: object) -> str:
    text = str(value)
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate or summarize Project1 expert-evaluation forms.")
    parser.add_argument("--package-dir", default="")
    parser.add_argument("--write-xlsx", action="store_true")
    parser.add_argument("--ratings-xlsx", default="")
    parser.add_argument("--ratings-dir", default="")
    parser.add_argument("--out-dir", default="results")
    args = parser.parse_args()
    if args.write_xlsx:
        if not args.package_dir:
            raise SystemExit("--write-xlsx requires --package-dir")
        print(write_expert_xlsx_forms(args.package_dir))
    if args.ratings_xlsx:
        print(json.dumps(summarize_expert_ratings(args.ratings_xlsx, args.out_dir), indent=2, ensure_ascii=False))
    if args.ratings_dir:
        print(json.dumps(summarize_expert_ratings_dir(args.ratings_dir, args.out_dir), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
