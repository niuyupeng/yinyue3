from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from chorale.expert_eval_tools import (
    ABSOLUTE_COLUMNS,
    ABSOLUTE_SHEET_NAMES,
    BACKGROUND_COLUMNS,
    PAIRED_COLUMNS,
    PAIRED_SHEET_NAMES,
    PREFERENCE_VALUES,
)

MIN_VALID_RATERS = 3
BACKGROUND_SHEET_NAMES = ["专家背景", "Background", "Rater background"]


def validate_returned_ratings(ratings_dir: str | Path) -> dict[str, object]:
    directory = Path(ratings_dir)
    files = sorted(
        path
        for path in directory.glob("*.xlsx")
        if not path.name.startswith("~$") and "template" not in path.name.lower()
    ) if directory.is_dir() else []
    file_reports = [validate_workbook(path) for path in files]
    valid_files = [item for item in file_reports if item["valid"]]
    absolute_rows = sum(int(item.get("absolute_completed_rows", 0) or 0) for item in file_reports)
    paired_rows = sum(int(item.get("paired_completed_rows", 0) or 0) for item in file_reports)
    rater_ids = [str(item.get("rater_id") or "").strip() for item in valid_files]
    duplicate_rater_ids = sorted({rater_id for rater_id in rater_ids if rater_ids.count(rater_id) > 1})
    release_gate_issues = build_release_gate_issues(
        valid_file_count=len(valid_files),
        absolute_rows=absolute_rows,
        paired_rows=paired_rows,
        rater_ids=rater_ids,
        duplicate_rater_ids=duplicate_rater_ids,
    )
    status = "ready_to_summarize" if not release_gate_issues else "expert evaluation pending"
    return {
        "schema": "project1_expert_return_intake_v1",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "ratings_dir": str(directory),
        "status": status,
        "ready_for_commercial_summary": status == "ready_to_summarize",
        "minimum_valid_raters": MIN_VALID_RATERS,
        "rating_file_count": len(files),
        "valid_rating_file_count": len(valid_files),
        "unique_rater_ids": sorted(set(rater_ids)),
        "duplicate_rater_ids": duplicate_rater_ids,
        "absolute_completed_rows": absolute_rows,
        "paired_completed_rows": paired_rows,
        "release_gate_issues": release_gate_issues,
        "files": file_reports,
        "next_step": (
            "Run scripts/summarize_project1_expert_ratings.ps1 after confirming the returned files are final."
            if status == "ready_to_summarize"
            else (
                "Collect at least three completed expert rating workbooks from distinct raters in "
                "expert_eval/project1/returned_ratings, then rerun validation."
            )
        ),
    }


def build_release_gate_issues(
    valid_file_count: int,
    absolute_rows: int,
    paired_rows: int,
    rater_ids: list[str],
    duplicate_rater_ids: list[str],
) -> list[str]:
    issues: list[str] = []
    if valid_file_count < MIN_VALID_RATERS:
        issues.append(f"need at least {MIN_VALID_RATERS} valid returned workbooks; found {valid_file_count}")
    if absolute_rows <= 0:
        issues.append("no completed absolute score-rating rows")
    if paired_rows <= 0:
        issues.append("no completed paired-comparison rows")
    if len(set(rater_ids)) < MIN_VALID_RATERS:
        issues.append(f"need at least {MIN_VALID_RATERS} distinct non-empty rater IDs; found {len(set(rater_ids))}")
    if duplicate_rater_ids:
        issues.append(f"duplicate rater IDs detected: {duplicate_rater_ids}")
    return issues


def validate_workbook(path: Path) -> dict[str, object]:
    issues: list[str] = []
    rater_id = ""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        return {
            "file": str(path),
            "valid": False,
            "issues": [f"cannot open workbook: {type(exc).__name__}: {exc}"],
            "rater_id": "",
            "absolute_completed_rows": 0,
            "paired_completed_rows": 0,
        }
    background_sheet = find_sheet(wb, BACKGROUND_SHEET_NAMES)
    absolute_sheet = find_sheet(wb, ABSOLUTE_SHEET_NAMES)
    paired_sheet = find_sheet(wb, PAIRED_SHEET_NAMES)
    if background_sheet is None:
        issues.append("missing rater background sheet")
    else:
        background_header = [cell.value for cell in background_sheet[1]]
        missing = [col for col in BACKGROUND_COLUMNS if col not in background_header]
        if missing:
            issues.append(f"background sheet missing columns: {missing}")
        rater_id = extract_rater_id(background_sheet)
        if not rater_id:
            issues.append("missing rater_id in background sheet")
    if absolute_sheet is None:
        issues.append("missing absolute rating sheet")
        absolute_rows = 0
    else:
        absolute_header = [cell.value for cell in absolute_sheet[1]]
        missing = [col for col in ABSOLUTE_COLUMNS if col not in absolute_header]
        if missing:
            issues.append(f"absolute sheet missing columns: {missing}")
        absolute_rows = count_completed_rows(absolute_sheet, ABSOLUTE_COLUMNS[2:10])
        validate_complete_rows(absolute_sheet, ABSOLUTE_COLUMNS[2:10], "absolute rating", issues)
        validate_numeric_range(absolute_sheet, ABSOLUTE_COLUMNS[2:10], 1, 5, "absolute rating", issues)
        if absolute_rows == 0:
            issues.append("no completed absolute rating rows detected")
    if paired_sheet is None:
        issues.append("missing paired comparison sheet")
        paired_rows = 0
    else:
        paired_header = [cell.value for cell in paired_sheet[1]]
        missing = [col for col in PAIRED_COLUMNS if col not in paired_header]
        if missing:
            issues.append(f"paired sheet missing columns: {missing}")
        paired_rows = count_completed_rows(paired_sheet, PAIRED_COLUMNS[3:11])
        validate_complete_rows(paired_sheet, PAIRED_COLUMNS[3:12], "paired comparison", issues)
        validate_preferences(paired_sheet, PAIRED_COLUMNS[3:11], issues)
        validate_numeric_range(paired_sheet, [PAIRED_COLUMNS[11]], 1, 5, "paired confidence", issues)
        if paired_rows == 0:
            issues.append("no completed paired comparison rows detected")
    return {
        "file": str(path),
        "valid": not issues,
        "issues": issues,
        "rater_id": rater_id,
        "absolute_completed_rows": absolute_rows,
        "paired_completed_rows": paired_rows,
        "sheet_names": wb.sheetnames,
    }


def find_sheet(wb, names: list[str]):
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return None


def count_completed_rows(sheet, metric_columns: list[str]) -> int:
    header = [cell.value for cell in sheet[1]]
    indices = [header.index(col) + 1 for col in metric_columns if col in header]
    count = 0
    for row_idx in range(2, sheet.max_row + 1):
        if any(sheet.cell(row_idx, col_idx).value not in (None, "") for col_idx in indices):
            count += 1
    return count


def extract_rater_id(sheet) -> str:
    header = [cell.value for cell in sheet[1]]
    if BACKGROUND_COLUMNS[0] in header:
        col_idx = header.index(BACKGROUND_COLUMNS[0]) + 1
    else:
        col_idx = 1
    for row_idx in range(2, sheet.max_row + 1):
        value = sheet.cell(row_idx, col_idx).value
        if value not in (None, ""):
            return str(value).strip()
    return ""


def validate_complete_rows(sheet, columns: list[str], label: str, issues: list[str]) -> None:
    header = [cell.value for cell in sheet[1]]
    indices = [(col, header.index(col) + 1) for col in columns if col in header]
    if not indices:
        return
    for row_idx in range(2, sheet.max_row + 1):
        values = [(col, sheet.cell(row_idx, col_idx).value) for col, col_idx in indices]
        if not any(value not in (None, "") for _, value in values):
            continue
        missing = [col for col, value in values if value in (None, "")]
        if missing:
            issues.append(f"{label} incomplete row in {sheet.title}!{row_idx}: missing {missing}")


def validate_numeric_range(
    sheet,
    columns: list[str],
    low: int,
    high: int,
    label: str,
    issues: list[str],
) -> None:
    header = [cell.value for cell in sheet[1]]
    for column in columns:
        if column not in header:
            continue
        col_idx = header.index(column) + 1
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            if value in (None, ""):
                continue
            if not isinstance(value, (int, float)) or not low <= float(value) <= high:
                issues.append(f"{label} out of range in {sheet.title}!{cell.coordinate}: {value!r}")


def validate_preferences(sheet, columns: list[str], issues: list[str]) -> None:
    header = [cell.value for cell in sheet[1]]
    allowed = set(PREFERENCE_VALUES)
    for column in columns:
        if column not in header:
            continue
        col_idx = header.index(column) + 1
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            if value in (None, ""):
                continue
            if str(value).strip() not in allowed:
                issues.append(f"paired preference invalid in {sheet.title}!{cell.coordinate}: {value!r}")


def write_outputs(report: dict[str, object], out_json: str | Path) -> dict[str, str]:
    out = Path(out_json)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    md = out.with_suffix(".md")
    md.write_text(make_markdown(report), encoding="utf-8")
    return {"json": str(out), "markdown": str(md)}


def make_markdown(report: dict[str, object]) -> str:
    lines = [
        "# Project1 Expert Return Intake Report",
        "",
        f"Status: **{report.get('status')}**",
        f"Rating files: {report.get('rating_file_count')}",
        f"Valid files: {report.get('valid_rating_file_count')}",
        f"Minimum valid raters: {report.get('minimum_valid_raters')}",
        f"Unique rater IDs: {', '.join(report.get('unique_rater_ids', []) or []) or 'none'}",
        f"Duplicate rater IDs: {', '.join(report.get('duplicate_rater_ids', []) or []) or 'none'}",
        f"Absolute completed rows: {report.get('absolute_completed_rows')}",
        f"Paired completed rows: {report.get('paired_completed_rows')}",
        "",
        "## Release Gate Issues",
        "",
    ]
    issues = report.get("release_gate_issues") or []
    if isinstance(issues, list) and issues:
        lines.extend(f"- {issue}" for issue in issues)
    else:
        lines.append("- none")
    lines.extend(["", "## Next Step", "", str(report.get("next_step", ""))])
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate returned Project1 expert rating workbooks before summarizing.")
    parser.add_argument("--ratings-dir", default="expert_eval/project1/returned_ratings")
    parser.add_argument("--out-json", default="results/project1_expert_return_intake_report_latest.json")
    args = parser.parse_args()
    report = validate_returned_ratings(args.ratings_dir)
    outputs = write_outputs(report, args.out_json)
    print(json.dumps({"report": report, "outputs": outputs}, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
