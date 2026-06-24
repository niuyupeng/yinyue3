from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from chorale.delivery_issue_debugger import debug_delivery_item, latest_package_dir


VALID_VARIANTS = {
    "full_choir",
    "piano_reference",
    "stem_soprano",
    "stem_alto",
    "stem_tenor",
    "stem_bass",
}

FIELD_ALIASES = {
    "issue_id": ["问题编号", "issue_id", "issue id", "id"],
    "score_id": ["谱例编号(score_id)", "谱例编号", "score_id", "score id"],
    "group": ["材料类型(absolute/paired)", "材料类型", "group", "material_type"],
    "variant": ["音频版本", "variant", "playback_variant"],
    "time_sec": ["问题时间点(秒)", "问题时间点", "time_sec", "time", "timestamp"],
    "category": ["问题类别", "category", "issue_category"],
    "severity": ["严重程度(1-5)", "严重程度", "severity"],
    "description": ["具体描述", "description", "issue_description"],
    "affects_rating": ["是否影响评分", "affects_rating", "affects score"],
    "reporter": ["反馈人", "reporter", "rater_id"],
    "notes": ["备注", "notes"],
}


def intake_review_issues(
    issues_path: str | Path = "expert_eval/project1/returned_issues",
    *,
    package_dir: str | Path | None = None,
) -> dict[str, Any]:
    issues_root = Path(issues_path)
    package = Path(package_dir) if package_dir else latest_package_dir(".")
    files = collect_issue_files(issues_root)
    rows: list[dict[str, Any]] = []
    source_file_errors: list[str] = []

    for file_path in files:
        try:
            raw_rows = read_issue_file(file_path)
        except Exception as exc:  # pragma: no cover - defensive CLI path
            source_file_errors.append(f"{file_path}: {type(exc).__name__}: {exc}")
            continue
        for index, raw in enumerate(raw_rows, start=2):
            normalized = normalize_issue_row(raw)
            if is_blank_or_example(normalized):
                continue
            rows.append(classify_issue_row(normalized, package, file_path, index))

    invalid_rows = [row for row in rows if row["row_status"] == "invalid"]
    matched_rows = [row for row in rows if row["manifest_match"] is True]
    unmatched_rows = [row for row in rows if row["row_status"] == "unmatched"]
    needs_attention = [
        row
        for row in rows
        if row.get("automatic_item_status") not in {"pass", "", None}
        or int(row.get("severity", 0) or 0) >= 4
    ]
    status = "no_issue_files" if not files else "ready_for_triage"
    if source_file_errors or invalid_rows:
        status = "has_invalid_rows"
    return {
        "schema": "project1_review_issue_intake_v1",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "issues_path": str(issues_root),
        "package_dir": str(package),
        "status": status,
        "issue_file_count": len(files),
        "source_files": [str(path) for path in files],
        "source_file_errors": source_file_errors,
        "accepted_issue_count": len(rows),
        "matched_issue_count": len(matched_rows),
        "unmatched_issue_count": len(unmatched_rows),
        "invalid_issue_count": len(invalid_rows),
        "needs_attention_count": len(needs_attention),
        "high_severity_count": sum(1 for row in rows if int(row.get("severity", 0) or 0) >= 4),
        "rows": rows,
        "next_action": next_action(files, invalid_rows, unmatched_rows),
    }


def collect_issue_files(path: Path) -> list[Path]:
    if path.is_file() and path.suffix.lower() in {".csv", ".xlsx"}:
        return [path]
    if not path.is_dir():
        return []
    return sorted(
        item
        for item in path.iterdir()
        if item.is_file()
        and item.suffix.lower() in {".csv", ".xlsx"}
        and not item.name.startswith("~$")
    )


def read_issue_file(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            return list(csv.DictReader(handle))
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_cell(value) for value in rows[0]]
    data_rows: list[dict[str, str]] = []
    for values in rows[1:]:
        data_rows.append({headers[i]: clean_cell(value) for i, value in enumerate(values) if i < len(headers)})
    return data_rows


def normalize_issue_row(row: dict[str, str]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    lowered = {normalize_header(key): value for key, value in row.items()}
    for canonical, aliases in FIELD_ALIASES.items():
        value = ""
        for alias in aliases:
            value = lowered.get(normalize_header(alias), "")
            if str(value).strip():
                break
        normalized[canonical] = clean_cell(value)
    return normalized


def classify_issue_row(row: dict[str, str], package: Path, source_file: Path, source_row: int) -> dict[str, Any]:
    validation_issues = validate_issue(row)
    result: dict[str, Any] = {
        "source_file": str(source_file),
        "source_row": source_row,
        "issue_id": row.get("issue_id", ""),
        "score_id": row.get("score_id", ""),
        "group": row.get("group", ""),
        "variant": row.get("variant", ""),
        "time_sec": row.get("time_sec", ""),
        "category": row.get("category", ""),
        "severity": parse_int(row.get("severity", "")),
        "description": row.get("description", ""),
        "affects_rating": row.get("affects_rating", ""),
        "reporter": row.get("reporter", ""),
        "notes": row.get("notes", ""),
        "validation_issues": validation_issues,
        "manifest_match": False,
        "automatic_item_status": "",
        "media_status": "",
        "conformance_status": "",
        "source_musicxml": "",
        "render_musicxml": "",
        "midi": "",
        "mp3": "",
        "debug_issues": [],
    }
    if validation_issues:
        result["row_status"] = "invalid"
        return result

    time_value = parse_optional_float(row.get("time_sec", ""))
    debug = debug_delivery_item(
        package,
        row["score_id"],
        row["variant"],
        time_sec=time_value,
    )
    result["automatic_item_status"] = str(debug.get("status", ""))
    if debug.get("status") == "not_found":
        result["row_status"] = "unmatched"
        result["debug_issues"] = [str(debug.get("message", "not found"))]
        return result

    manifest_row = debug.get("manifest_row", {})
    if isinstance(manifest_row, dict):
        result["group"] = result["group"] or str(manifest_row.get("group", ""))
        for key in ["source_musicxml", "render_musicxml", "midi", "mp3"]:
            result[key] = str(manifest_row.get(key, ""))
    media = debug.get("media_audit", {})
    conformance = debug.get("conformance_audit", {})
    if isinstance(media, dict):
        result["media_status"] = str(media.get("status", ""))
    if isinstance(conformance, dict):
        result["conformance_status"] = str(conformance.get("status", ""))
    debug_issues = debug.get("issues", [])
    result["debug_issues"] = debug_issues if isinstance(debug_issues, list) else []
    timepoint = debug.get("timepoint_diagnostic", {})
    if isinstance(timepoint, dict) and timepoint:
        result["timepoint_diagnostic"] = timepoint
        result["estimated_measure"] = timepoint.get("estimated_measure", "")
        result["estimated_beat"] = timepoint.get("estimated_beat", "")
        result["measure_relative_offset_quarter"] = timepoint.get("measure_relative_offset_quarter", "")
        result["measure_duration_quarter"] = timepoint.get("measure_duration_quarter", "")
        result["time_signature"] = timepoint.get("time_signature", "")
        result["render_nearby_pitches"] = summarize_nearby_pitches(timepoint.get("render_notes_near_time", []))
        result["source_nearby_pitches"] = summarize_nearby_pitches(timepoint.get("source_notes_near_time", []))
        issues = timepoint.get("issues", [])
        result["timepoint_issues"] = issues if isinstance(issues, list) else []
    result["manifest_match"] = True
    result["row_status"] = "matched"
    return result


def validate_issue(row: dict[str, str]) -> list[str]:
    issues: list[str] = []
    if not row.get("score_id"):
        issues.append("missing score_id")
    variant = row.get("variant", "")
    if not variant:
        issues.append("missing variant")
    elif variant not in VALID_VARIANTS:
        issues.append(f"invalid variant: {variant}")
    severity_raw = row.get("severity", "")
    severity = parse_int(severity_raw)
    if severity_raw and severity not in {1, 2, 3, 4, 5}:
        issues.append(f"severity must be 1-5: {severity_raw}")
    if not row.get("description"):
        issues.append("missing description")
    return issues


def is_blank_or_example(row: dict[str, str]) -> bool:
    values = [str(value).strip() for value in row.values()]
    if not any(values):
        return True
    issue_id = row.get("issue_id", "").strip().upper()
    score_id = row.get("score_id", "").strip().upper()
    return issue_id == "EXAMPLE" or score_id in {"P1S??", "P1P??", "P1SXX"}


def parse_int(value: str) -> int:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


def parse_optional_float(value: str) -> float | None:
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return None
    return parsed if parsed >= 0 else None


def summarize_nearby_pitches(rows: object) -> str:
    if not isinstance(rows, list):
        return ""
    tokens: list[str] = []
    for row in rows[:16]:
        if not isinstance(row, dict):
            continue
        pitches = row.get("pitches", [])
        pitch_text = ",".join(str(pitch) for pitch in pitches) if isinstance(pitches, list) else str(pitches)
        if pitch_text:
            tokens.append(f"{row.get('part_name')}@{row.get('offset_quarter')}:{pitch_text}")
    return "; ".join(tokens)


def clean_cell(value: object) -> str:
    return "" if value is None else " ".join(str(value).strip().split())


def normalize_header(value: str) -> str:
    return clean_cell(value).lower().replace(" ", "").replace("_", "")


def next_action(files: list[Path], invalid_rows: list[dict[str, Any]], unmatched_rows: list[dict[str, Any]]) -> str:
    if not files:
        return "No returned issue files found. Put reviewer issue CSV/XLSX files under expert_eval/project1/returned_issues."
    if invalid_rows:
        return "Fix invalid rows before using the issue report for commercial triage."
    if unmatched_rows:
        return "Check unmatched score_id/variant values against the current delivery ZIP manifest."
    return "Use rows with severity >= 4 or automatic_item_status != pass as the first triage queue."


def write_outputs(report: dict[str, Any], out_json: str | Path) -> dict[str, str]:
    out = Path(out_json)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    out_csv = out.with_suffix(".csv")
    write_rows_csv(report, out_csv)
    out_md = out.with_suffix(".md")
    out_md.write_text(make_markdown(report), encoding="utf-8")
    return {"json": str(out), "csv": str(out_csv), "markdown": str(out_md)}


def write_rows_csv(report: dict[str, Any], out_csv: Path) -> None:
    rows = report.get("rows", [])
    fieldnames = [
        "row_status",
        "issue_id",
        "score_id",
        "group",
        "variant",
        "time_sec",
        "category",
        "severity",
        "description",
        "affects_rating",
        "reporter",
        "manifest_match",
        "automatic_item_status",
        "media_status",
        "conformance_status",
        "estimated_measure",
        "estimated_beat",
        "measure_relative_offset_quarter",
        "measure_duration_quarter",
        "time_signature",
        "render_nearby_pitches",
        "source_nearby_pitches",
        "timepoint_issues",
        "source_musicxml",
        "render_musicxml",
        "midi",
        "mp3",
        "validation_issues",
        "debug_issues",
    ]
    with out_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows if isinstance(rows, list) else []:
            if not isinstance(row, dict):
                continue
            flat = {key: row.get(key, "") for key in fieldnames}
            for key in ["validation_issues", "debug_issues", "timepoint_issues"]:
                value = flat.get(key, "")
                flat[key] = "; ".join(str(item) for item in value) if isinstance(value, list) else str(value)
            writer.writerow(flat)


def make_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# Project1 Review Issue Intake Report",
        "",
        f"Status: **{report.get('status')}**",
        f"Package: `{report.get('package_dir')}`",
        f"Issue files: {report.get('issue_file_count')}",
        f"Accepted issues: {report.get('accepted_issue_count')}",
        f"Matched issues: {report.get('matched_issue_count')}",
        f"Invalid issues: {report.get('invalid_issue_count')}",
        f"Needs attention: {report.get('needs_attention_count')}",
        "",
        str(report.get("next_action", "")),
        "",
    ]
    rows = report.get("rows", [])
    if isinstance(rows, list) and rows:
        lines.extend(["## Triage Rows", ""])
        for row in rows[:100]:
            if not isinstance(row, dict):
                continue
            lines.append(
                f"- `{row.get('row_status')}` {row.get('score_id')}/{row.get('variant')} "
                f"severity={row.get('severity')} status={row.get('automatic_item_status')} "
                f"m={row.get('estimated_measure', '')} beat={row.get('estimated_beat', '')}: "
                f"{row.get('description')}"
            )
    else:
        lines.append("No reviewer issue rows were ingested.")
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest returned Project1 reviewer issue reports.")
    parser.add_argument("--issues-path", default="expert_eval/project1/returned_issues")
    parser.add_argument("--package-dir", default="")
    parser.add_argument("--out-json", default="results/project1_review_issue_intake_latest.json")
    parser.add_argument("--strict", action="store_true", help="Fail on invalid or unmatched returned issue rows.")
    args = parser.parse_args()
    package = Path(args.package_dir) if args.package_dir else latest_package_dir(".")
    report = intake_review_issues(args.issues_path, package_dir=package)
    outputs = write_outputs(report, args.out_json)
    print(json.dumps({"report": {k: v for k, v in report.items() if k != "rows"}, "outputs": outputs}, indent=2, ensure_ascii=False))
    if args.strict and (report["invalid_issue_count"] or report["unmatched_issue_count"]):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
