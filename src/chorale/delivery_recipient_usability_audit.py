from __future__ import annotations

import argparse
import csv
import io
import json
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook


REQUIRED_TOP_LEVEL = {
    "START_HERE_CN.html",
    "START_HERE_CN.md",
    "score_audio_player.html",
    "OPEN_PROJECT1_REVIEW_PACKAGE.ps1",
    "PROJECT1_PACKAGE_SELF_TEST.ps1",
    "PROJECT1_PACKAGE_SELF_TEST_README_CN.md",
    "VERIFY_DELIVERY_INTEGRITY.ps1",
    "VERIFY_DELIVERY_INTEGRITY_README_CN.md",
    "DELIVERY_README_CN.md",
    "README_CN.md",
    "README_FOR_EXPERTS.md",
    "RETURN_FILES_CHECKLIST.md",
    "REVIEW_ISSUE_REPORT_TEMPLATE.csv",
    "REVIEW_ISSUE_REPORT_GUIDE_CN.md",
    "SCORE_AUDIO_CORRESPONDENCE.csv",
    "SCORE_AUDIO_CORRESPONDENCE_README.md",
    "THIRD_PARTY_PLAYBACK_NOTICES.md",
}

REQUIRED_FORMS = {
    "forms/project1_expert_rating_forms_CN.xlsx",
    "forms/rater_background_form_project1_CN.csv",
    "forms/absolute_rating_form_project1_CN.csv",
    "forms/paired_comparison_form_project1_CN.csv",
    "forms/中文填表说明.md",
    "forms/评分参考与示例_CN.md",
}

EXPECTED_WORKBOOK_HEADERS = {
    "评分说明": ["分数", "含义"],
    "专家背景": [
        "评分人编号",
        "最高音乐训练",
        "主要专长",
        "和声对位经验年数",
        "合唱或作曲教学经验年数",
        "是否经常使用传统功能和声_yes_no",
        "备注",
    ],
    "逐首评分": [
        "score_id",
        "乐谱文件名",
        "和声正确性_1到5",
        "声部进行与对位正确性_1到5",
        "七和弦解决正确性_1到5",
        "终止式质量_1到5",
        "SATB可唱性_1到5",
        "传统众赞歌风格一致性_1到5",
        "作曲和声教学用途_1到5",
        "整体质量_1到5",
        "发现的主要问题",
        "备注",
    ],
    "AB配对比较": [
        "pair_id",
        "版本A文件",
        "版本B文件",
        "和声方面更好者",
        "声部进行方面更好者",
        "七和弦解决方面更好者",
        "终止式方面更好者",
        "SATB可唱性更好者",
        "风格一致性更好者",
        "教学用途更好者",
        "总体更好者",
        "判断信心_1到5",
        "备注",
    ],
}

ISSUE_TEMPLATE_HEADERS = [
    "问题编号",
    "谱例编号(score_id)",
    "材料类型(absolute/paired)",
    "音频版本",
    "问题时间点(秒)",
    "问题类别",
    "严重程度(1-5)",
    "具体描述",
    "是否影响评分",
    "反馈人",
    "备注",
]

CORRESPONDENCE_REQUIRED_COLUMNS = {
    "group",
    "score_id",
    "variant",
    "source_musicxml",
    "render_musicxml",
    "midi",
    "mp3",
    "status",
}
REQUIRED_PLAYBACK_VARIANTS = {
    "full_choir",
    "piano_reference",
    "stem_soprano",
    "stem_alto",
    "stem_tenor",
    "stem_bass",
}
EXPECTED_CORRESPONDENCE_ROWS = 240

BAD_TEXT_MARKERS = ["锟", "鈥", "â", "�", "浜", "涔", "璇", "鎵", "闂", "鐨", "鍙", "涓", "瀹", "鏂"]
USER_FACING_TEXT_FILES = [
    "START_HERE_CN.html",
    "START_HERE_CN.md",
    "DELIVERY_README_CN.md",
    "README_CN.md",
    "README_FOR_EXPERTS.md",
    "RETURN_FILES_CHECKLIST.md",
    "REVIEW_ISSUE_REPORT_GUIDE_CN.md",
    "SCORE_AUDIO_CORRESPONDENCE_README.md",
    "forms/中文填表说明.md",
    "forms/评分参考与示例_CN.md",
]


class Reader(Protocol):
    source: str

    def exists(self, path: str) -> bool: ...

    def read_text(self, path: str) -> str: ...

    def read_bytes(self, path: str) -> bytes: ...

    def list_files(self) -> list[str]: ...


@dataclass
class DirReader:
    root: Path

    @property
    def source(self) -> str:
        return str(self.root)

    def exists(self, path: str) -> bool:
        return (self.root / path).is_file()

    def read_text(self, path: str) -> str:
        return (self.root / path).read_text(encoding="utf-8-sig", errors="replace")

    def read_bytes(self, path: str) -> bytes:
        return (self.root / path).read_bytes()

    def list_files(self) -> list[str]:
        return sorted(item.relative_to(self.root).as_posix() for item in self.root.rglob("*") if item.is_file())


@dataclass
class ZipReader:
    zip_path: Path

    def __post_init__(self) -> None:
        with zipfile.ZipFile(self.zip_path) as archive:
            names = [item.filename.replace("\\", "/") for item in archive.infolist() if not item.is_dir()]
        roots = {name.split("/", 1)[0] for name in names if "/" in name}
        self._prefix = next(iter(roots)) + "/" if len(roots) == 1 and all(name.startswith(next(iter(roots)) + "/") for name in names) else ""
        self._names = {self._strip(name) for name in names}

    @property
    def source(self) -> str:
        return str(self.zip_path)

    def _strip(self, name: str) -> str:
        return name[len(self._prefix) :] if self._prefix and name.startswith(self._prefix) else name

    def _archive_name(self, path: str) -> str:
        return self._prefix + path.replace("\\", "/")

    def exists(self, path: str) -> bool:
        return path.replace("\\", "/") in self._names

    def read_text(self, path: str) -> str:
        return self.read_bytes(path).decode("utf-8-sig", errors="replace")

    def read_bytes(self, path: str) -> bytes:
        with zipfile.ZipFile(self.zip_path) as archive:
            return archive.read(self._archive_name(path))

    def list_files(self) -> list[str]:
        return sorted(self._names)


def audit_recipient_usability(package_dir: str | Path | None = None, zip_file: str | Path | None = None) -> dict[str, object]:
    if bool(package_dir) == bool(zip_file):
        raise ValueError("Provide exactly one of package_dir or zip_file.")
    reader: Reader = DirReader(Path(package_dir)) if package_dir else ZipReader(Path(zip_file))  # type: ignore[arg-type]
    issues: list[str] = []
    warnings: list[str] = []

    files = set(reader.list_files())
    for path in sorted(REQUIRED_TOP_LEVEL | REQUIRED_FORMS):
        if path not in files:
            issues.append(f"missing recipient-facing file: {path}")

    audit_start_page(reader, issues)
    audit_user_facing_text(reader, issues)
    workbook_summary = audit_workbook(reader, issues)
    issue_template_summary = audit_issue_template(reader, issues)
    correspondence_summary = audit_correspondence(reader, issues)
    audit_return_checklist(reader, issues)
    audit_open_scripts(reader, issues)
    audit_optional_legacy_forms(files, warnings)

    score = 100 if not issues else max(0, 100 - len(issues) * 10 - len(warnings) * 2)
    return {
        "schema": "project1_recipient_usability_audit_v1",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": reader.source,
        "recipient_usability_score": score,
        "all_pass": not issues,
        "status": "pass" if not issues else "failed",
        "file_count": len(files),
        "workbook": workbook_summary,
        "issue_template": issue_template_summary,
        "score_audio_correspondence": correspondence_summary,
        "issues": issues,
        "warnings": warnings,
    }


def audit_start_page(reader: Reader, issues: list[str]) -> None:
    if not reader.exists("START_HERE_CN.html"):
        return
    text = reader.read_text("START_HERE_CN.html")
    required = [
        "交付包入口",
        "score_audio_player.html",
        "forms/project1_expert_rating_forms_CN.xlsx",
        "MusicXML",
        "不是神经音频生成",
    ]
    for snippet in required:
        if snippet not in text:
            issues.append(f"START_HERE_CN.html missing readable guidance: {snippet}")


def audit_user_facing_text(reader: Reader, issues: list[str]) -> None:
    for path in USER_FACING_TEXT_FILES:
        if not reader.exists(path):
            continue
        text = reader.read_text(path)
        marker_count = sum(text.count(marker) for marker in BAD_TEXT_MARKERS)
        if marker_count >= 3:
            issues.append(f"user-facing text may contain mojibake: {path} markers={marker_count}")


def audit_workbook(reader: Reader, issues: list[str]) -> dict[str, object]:
    path = "forms/project1_expert_rating_forms_CN.xlsx"
    if not reader.exists(path):
        return {"status": "missing", "sheets": []}
    try:
        workbook = load_workbook(io.BytesIO(reader.read_bytes(path)), read_only=True, data_only=True)
    except Exception as exc:
        issues.append(f"could not open expert workbook: {type(exc).__name__}: {exc}")
        return {"status": "parse_failed", "sheets": []}
    sheet_names = workbook.sheetnames
    for sheet, expected_headers in EXPECTED_WORKBOOK_HEADERS.items():
        if sheet not in workbook.sheetnames:
            issues.append(f"expert workbook missing sheet: {sheet}")
            continue
        ws = workbook[sheet]
        headers = [clean_cell(ws.cell(1, col).value) for col in range(1, len(expected_headers) + 1)]
        if headers != expected_headers:
            issues.append(f"expert workbook sheet {sheet} headers mismatch: {headers}")
    return {"status": "pass", "sheets": sheet_names}


def audit_issue_template(reader: Reader, issues: list[str]) -> dict[str, object]:
    path = "REVIEW_ISSUE_REPORT_TEMPLATE.csv"
    if not reader.exists(path):
        return {"status": "missing", "headers": []}
    rows = list(csv.reader(io.StringIO(reader.read_text(path))))
    headers = rows[0] if rows else []
    if headers != ISSUE_TEMPLATE_HEADERS:
        issues.append(f"issue report template headers mismatch: {headers}")
    return {"status": "pass" if headers == ISSUE_TEMPLATE_HEADERS else "failed", "headers": headers, "example_rows": max(0, len(rows) - 1)}


def audit_correspondence(reader: Reader, issues: list[str]) -> dict[str, object]:
    path = "SCORE_AUDIO_CORRESPONDENCE.csv"
    if not reader.exists(path):
        return {"status": "missing", "rows": 0}
    rows = list(csv.DictReader(io.StringIO(reader.read_text(path))))
    headers = set(rows[0].keys()) if rows else set()
    missing_cols = sorted(CORRESPONDENCE_REQUIRED_COLUMNS - headers)
    if missing_cols:
        issues.append(f"SCORE_AUDIO_CORRESPONDENCE.csv missing columns: {missing_cols}")
    if len(rows) != EXPECTED_CORRESPONDENCE_ROWS:
        issues.append(
            f"SCORE_AUDIO_CORRESPONDENCE.csv row count {len(rows)}, "
            f"expected {EXPECTED_CORRESPONDENCE_ROWS}"
        )
    missing_refs = []
    status_errors = []
    variants_by_score: dict[tuple[str, str], set[str]] = {}
    for row in rows:
        group = str(row.get("group", "")).strip()
        score_id = str(row.get("score_id", "")).strip()
        variant = str(row.get("variant", "")).strip()
        variants_by_score.setdefault((group, score_id), set()).add(variant)
        if row.get("status") != "ok":
            status_errors.append(f"{group}/{score_id}/{variant}: status={row.get('status')!r}")
        for key in ["source_musicxml", "render_musicxml", "midi", "mp3"]:
            rel = str(row.get(key, "")).replace("\\", "/")
            if not rel:
                missing_refs.append(f"{group}/{score_id}/{variant}: empty {key}")
            elif not reader.exists(rel):
                missing_refs.append(f"{group}/{score_id}/{variant}: {key} -> {rel}")
        wav_rel = str(row.get("wav", "")).replace("\\", "/")
        if wav_rel and not reader.exists(wav_rel):
            missing_refs.append(f"{group}/{score_id}/{variant}: wav -> {wav_rel}")
    if missing_refs:
        issues.append("SCORE_AUDIO_CORRESPONDENCE.csv has missing references: " + "; ".join(missing_refs[:10]))
    if status_errors:
        issues.append("SCORE_AUDIO_CORRESPONDENCE.csv has non-ok statuses: " + "; ".join(status_errors[:10]))
    variant_errors = []
    for (group, score_id), variants in sorted(variants_by_score.items()):
        missing = REQUIRED_PLAYBACK_VARIANTS - variants
        extra = variants - REQUIRED_PLAYBACK_VARIANTS
        if missing:
            variant_errors.append(f"{group}/{score_id}: missing variants {','.join(sorted(missing))}")
        if extra:
            variant_errors.append(f"{group}/{score_id}: unexpected variants {','.join(sorted(extra))}")
    if variant_errors:
        issues.append("SCORE_AUDIO_CORRESPONDENCE.csv variant coverage errors: " + "; ".join(variant_errors[:10]))
    passed = (
        not missing_cols
        and len(rows) == EXPECTED_CORRESPONDENCE_ROWS
        and not missing_refs
        and not status_errors
        and not variant_errors
    )
    return {
        "status": "pass" if passed else "failed",
        "rows": len(rows),
        "score_variant_groups": len(variants_by_score),
    }


def audit_return_checklist(reader: Reader, issues: list[str]) -> None:
    if not reader.exists("RETURN_FILES_CHECKLIST.md"):
        return
    text = reader.read_text("RETURN_FILES_CHECKLIST.md")
    for snippet in ["project1_expert_rating_forms_CN.xlsx", "rater_background_form_project1_CN.csv", "absolute_rating_form_project1_CN.csv", "paired_comparison_form_project1_CN.csv"]:
        if snippet not in text:
            issues.append(f"RETURN_FILES_CHECKLIST.md missing return item: {snippet}")


def audit_open_scripts(reader: Reader, issues: list[str]) -> None:
    for path, snippets in {
        "OPEN_PROJECT1_REVIEW_PACKAGE.ps1": ["PROJECT1_PACKAGE_SELF_TEST.ps1", "START_HERE_CN.html", "score_audio_player.html"],
        "PROJECT1_PACKAGE_SELF_TEST.ps1": ["project1_recipient_package_self_test_v1", "Project1 package self-test"],
        "VERIFY_DELIVERY_INTEGRITY.ps1": ["DELIVERY_FILE_MANIFEST", "SHA256"],
    }.items():
        if not reader.exists(path):
            continue
        text = reader.read_text(path)
        for snippet in snippets:
            if snippet not in text:
                issues.append(f"{path} missing expected snippet: {snippet}")


def audit_optional_legacy_forms(files: set[str], warnings: list[str]) -> None:
    legacy = sorted(path for path in files if path.startswith("forms/") and path.endswith(".csv") and not path.endswith("_CN.csv"))
    if legacy:
        warnings.append("legacy English CSV alternatives are present; START_HERE should direct reviewers to the CN workbook first")


def clean_cell(value: object) -> str:
    return "" if value is None else " ".join(str(value).strip().split())


def write_outputs(summary: dict[str, object], out_json: str | Path) -> dict[str, str]:
    out = Path(out_json)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    out_md = out.with_suffix(".md")
    out_md.write_text(make_markdown(summary), encoding="utf-8")
    return {"json": str(out), "markdown": str(out_md)}


def make_markdown(summary: dict[str, object]) -> str:
    lines = [
        "# Project1 Recipient Usability Audit",
        "",
        f"Source: `{summary.get('source')}`",
        f"Score: **{summary.get('recipient_usability_score')}/100**",
        f"Status: **{summary.get('status')}**",
        "",
        "## Checks",
        "",
        f"- Workbook: `{summary.get('workbook')}`",
        f"- Issue template: `{summary.get('issue_template')}`",
        f"- Score-audio correspondence: `{summary.get('score_audio_correspondence')}`",
        "",
    ]
    issues = summary.get("issues", [])
    lines.extend(["## Issues", ""])
    if isinstance(issues, list) and issues:
        lines.extend(f"- {issue}" for issue in issues)
    else:
        lines.append("No recipient-usability issues detected.")
    warnings = summary.get("warnings", [])
    lines.extend(["", "## Warnings", ""])
    if isinstance(warnings, list) and warnings:
        lines.extend(f"- {warning}" for warning in warnings)
    else:
        lines.append("No warnings.")
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit Project1 recipient-facing package usability.")
    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--package-dir")
    src.add_argument("--zip-file")
    parser.add_argument("--out-json", default="results/project1_recipient_usability_audit_latest.json")
    args = parser.parse_args()
    summary = audit_recipient_usability(args.package_dir, args.zip_file)
    outputs = write_outputs(summary, args.out_json)
    print(json.dumps({"summary": summary, "outputs": outputs}, indent=2, ensure_ascii=False))
    if not summary["all_pass"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
